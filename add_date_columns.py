"""
One-time script: inserts three date-tracking columns between Amount (G) and
Check/Wire Reference (H) in QCD_Tracker_2026.xlsx.

New columns:
  H — Date Check Requested from IRA
  I — Date Check Received from IRA
  J — Date Check Mailed to Charity

Run once, then this script can be kept for reference or discarded.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SCRIPT_DIR = Path(__file__).parent
SHEET_NAME = "QCD Transaction Log"

HEADER_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL  = PatternFill(fill_type="solid", fgColor="FF2E75B6")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT       = Font(name="Arial", size=10)
BODY_FILL_GRAY  = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
BODY_FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
BODY_ALIGN      = Alignment(horizontal="center", vertical="top", wrap_text=True)

NEW_COLUMNS = [
    ("Date Check\nRequested\nfrom IRA",  "H"),
    ("Date Check\nReceived\nfrom IRA",   "I"),
    ("Date Check\nMailed to\nCharity",   "J"),
]


def _thin_border():
    s = Side(style="thin", color="FFBFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def main():
    matches = list(SCRIPT_DIR.glob("QCD_Tracker*.xlsx"))
    if not matches:
        print("ERROR: No QCD_Tracker*.xlsx found in this folder.")
        print(f"  Folder checked: {SCRIPT_DIR}")
        return

    xlsx_path = matches[0]

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except PermissionError:
        print(f"ERROR: {xlsx_path.name} is open in another program.")
        print("  Please close it and run this script again.")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found in {xlsx_path.name}.")
        return

    ws = wb[SHEET_NAME]

    # Safety check: if H3 already contains a date column header, don't insert again
    h3 = ws.cell(row=3, column=8).value
    if h3 and "requested" in str(h3).lower():
        print("Date columns already appear to be present. Nothing changed.")
        return

    # Verify G is Amount before inserting (confirms modify_spreadsheet.py has been run)
    g3 = ws.cell(row=3, column=7).value
    if g3 is None or "amount" not in str(g3).lower():
        print("ERROR: Column G does not appear to be 'Amount'.")
        print(f"  Found: {g3!r}")
        print("  Please run modify_spreadsheet.py first, then re-run this script.")
        return

    # Insert 3 blank columns at position 8 (H), shifting Check/Wire and beyond right
    ws.insert_cols(8, amount=3)

    border = _thin_border()
    for col_offset, (header_text, col_letter) in enumerate(NEW_COLUMNS):
        col_idx = 8 + col_offset

        # Header (row 3)
        h = ws.cell(row=3, column=col_idx)
        h.value     = header_text
        h.font      = HEADER_FONT
        h.fill      = HEADER_FILL
        h.alignment = HEADER_ALIGN

        # Body cells (rows 4-23) — even rows gray, odd rows white
        for row in range(4, 24):
            c = ws.cell(row=row, column=col_idx)
            c.font      = BODY_FONT
            c.fill      = BODY_FILL_GRAY if row % 2 == 0 else BODY_FILL_WHITE
            c.border    = border
            c.alignment = BODY_ALIGN

        ws.column_dimensions[col_letter].width = 13

    wb.save(xlsx_path)
    print(f"Done. {xlsx_path.name} updated.")
    print("  New columns inserted between Amount (G) and Check/Wire Reference (now K):")
    print("    H — Date Check Requested from IRA")
    print("    I — Date Check Received from IRA")
    print("    J — Date Check Mailed to Charity")
    print("  Open the spreadsheet in Excel to verify before use.")


if __name__ == "__main__":
    main()
