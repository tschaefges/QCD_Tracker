# Build command (run from this folder):
#   pyinstaller --onefile --console --name lookup_charity lookup_charity.py
# Or use build.bat

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl

# ---------------------------------------------------------------------------
# Path resolution — works both as .py and as a compiled PyInstaller EXE
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

SHEET_NAME  = "QCD Transaction Log"
DIVIDER     = "─" * 45
MAX_RESULTS = 10
FIRST_DATA_ROW = 4
LAST_DATA_ROW  = 23


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def read_config():
    config_path = BASE_DIR / "config.txt"
    if not config_path.exists() or not config_path.read_text().strip():
        print("ERROR: config.txt not found in this folder.")
        print("Please create a file named config.txt containing one line:")
        print("the UNC path to the shared IRS data folder.")
        print(r"Example:  \\House\Documents\IRS_Data")
        input("\nPress Enter to exit.")
        sys.exit(1)
    return config_path.read_text().strip()


def _try_read_csv(path):
    try:
        return pd.read_csv(path, dtype={"EIN": str, "ZIP": str}, encoding="utf-8")
    except Exception as exc:
        print(f"WARNING: Could not read {path.name}: {exc}")
        return None


def load_data(server_path):
    server_file = Path(server_path) / "eo_bmf.csv"
    backup_file = BASE_DIR / "eo_bmf_backup.csv"

    df = None
    source_path = None

    if server_file.exists():
        df = _try_read_csv(server_file)
        if df is not None:
            source_path = server_file
        elif backup_file.exists():
            mod_date = datetime.fromtimestamp(backup_file.stat().st_mtime).strftime("%Y-%m-%d")
            print(f"WARNING: Server file exists but could not be read. Falling back to local backup.")
            print(f"Backup last updated: {mod_date}")
            print("Data may not be current.")
            print()
            df = _try_read_csv(backup_file)
            if df is not None:
                source_path = backup_file

    elif backup_file.exists():
        mod_date = datetime.fromtimestamp(backup_file.stat().st_mtime).strftime("%Y-%m-%d")
        print(f"WARNING: Could not reach server at {server_path}")
        print("Using local backup copy instead.")
        print(f"Backup last updated: {mod_date}")
        print("Data may not be current.")
        print()
        df = _try_read_csv(backup_file)
        if df is not None:
            source_path = backup_file

    if source_path is None:
        print("ERROR: Could not find or read IRS data file.")
        print(f"Tried: {server_file}")
        print(f"Also tried: {backup_file}")
        print("Please check that the server is running and try again,")
        print("or run refresh_bmf.py to create a local backup.")
        input("\nPress Enter to exit.")
        sys.exit(1)

    data_date = datetime.fromtimestamp(source_path.stat().st_mtime).strftime("%Y-%m-%d")
    return df, data_date


def find_spreadsheet():
    matches = [
        p for p in BASE_DIR.glob("*.xlsx")
        if "qcd_tracker" in p.name.lower() and not p.name.startswith("~$")
    ]
    if not matches:
        print("ERROR: No spreadsheet found in this folder.")
        print("Looking for: any .xlsx file with \"QCD_Tracker\" in the name.")
        print(f"Folder checked: {BASE_DIR}")
        input("\nPress Enter to exit.")
        sys.exit(1)
    if len(matches) == 1:
        return matches[0]
    print("Multiple QCD_Tracker spreadsheets found:")
    for i, p in enumerate(matches, 1):
        print(f"  {i}. {p.name}")
    while True:
        choice = input("Enter number to select: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(matches):
            return matches[int(choice) - 1]


def format_ein(ein):
    s = str(ein).zfill(9)
    return f"{s[:2]}-{s[2:]}"


def format_zip(z):
    z = str(z).strip()
    if z.endswith("-0000"):
        return z[:5]
    return z


def format_address(row):
    street = str(row["STREET"]).strip() if pd.notna(row["STREET"]) else ""
    city   = str(row["CITY"]).strip()   if pd.notna(row["CITY"])   else ""
    state  = str(row["STATE"]).strip()  if pd.notna(row["STATE"])  else ""
    z      = format_zip(row["ZIP"])     if pd.notna(row["ZIP"])    else ""
    parts  = [p for p in [street, city] if p]
    location = f"{', '.join(parts)}, {state} {z}".strip(", ")
    return location


def find_next_row(xlsx_path):
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    except PermissionError:
        print(f"\nERROR: {xlsx_path.name} is open in another program.")
        print("Please close it and try again.")
        return None
    try:
        if SHEET_NAME not in wb.sheetnames:
            print(f"ERROR: Sheet '{SHEET_NAME}' not found in {xlsx_path.name}.")
            return None
        ws = wb[SHEET_NAME]
        for row_idx in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            if ws.cell(row=row_idx, column=4).value is None:
                return row_idx
        return None  # all rows full
    finally:
        wb.close()


def write_to_spreadsheet(xlsx_path, row_idx, name, ein_formatted, address):
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except PermissionError:
        print(f"\nERROR: {xlsx_path.name} is open in another program.")
        print("Please close it and try again.")
        return False

    try:
        if SHEET_NAME not in wb.sheetnames:
            print(f"ERROR: Sheet '{SHEET_NAME}' not found.")
            return False

        ws = wb[SHEET_NAME]

        # Verify column F is Charity Address before writing
        f3 = ws.cell(row=3, column=6).value
        if f3 is None or "address" not in str(f3).lower():
            print("\nERROR: Column F does not appear to be 'Charity Address'.")
            print(f"  Found: {f3!r}")
            print("  Please run modify_spreadsheet.py first to prepare the spreadsheet.")
            return False

        ws.cell(row=row_idx, column=4).value = name
        ws.cell(row=row_idx, column=5).value = ein_formatted
        ws.cell(row=row_idx, column=6).value = address

        try:
            wb.save(xlsx_path)
        except PermissionError:
            print(f"\nERROR: Could not save {xlsx_path.name}.")
            print("Please close it in Excel and try again.")
            return False
        except OSError as exc:
            print(f"\nERROR: Could not save {xlsx_path.name}: {exc}")
            return False

        return True
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------

def show_results(matches, term):
    total = len(matches)
    display = matches.iloc[:MAX_RESULTS]
    print(f"\nFound {total} match{'es' if total != 1 else ''} for \"{term}\":")
    if total > MAX_RESULTS:
        print(f"Showing {MAX_RESULTS} of {total} matches. Try a more specific search term for better results.")
    print()
    name_w = 34
    print(f"  {'#':<4} {'Name':<{name_w}}  {'EIN':<11}  City, State")
    print(f"  {'─':<4} {'─'*name_w}  {'─'*11}  {'─'*20}")
    for i, (_, row) in enumerate(display.iterrows(), 1):
        name_trunc = str(row["NAME"])[:name_w]
        city_state = f"{row['CITY']}, {row['STATE']}"
        print(f"  {i:<4} {name_trunc:<{name_w}}  {format_ein(row['EIN']):<11}  {city_state}")
    print()


def show_confirmation(row, next_row, xlsx_name):
    name    = str(row["NAME"])
    ein_fmt = format_ein(row["EIN"])
    address = format_address(row)

    print()
    print(DIVIDER)
    print("  Selected organization:")
    print()
    print(f"  Name:    {name}")
    print(f"  EIN:     {ein_fmt}")
    print(f"  Address: {address}")
    print()
    print(f"  Will write to row {next_row} of {xlsx_name}")
    print()
    print("  Columns to be filled:")
    print(f"    D (Charity Name)    → {name}")
    print(f"    E (Charity EIN)     → {ein_fmt}")
    print(f"    F (Charity Address) → {address}")
    print(DIVIDER)

    return name, ein_fmt, address


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    server_path = read_config()
    df, data_date = load_data(server_path)
    xlsx_path = find_spreadsheet()

    print(DIVIDER)
    print(" QCD Charity Lookup — IRS EO BMF Search")
    print(f" {len(df):,} organizations loaded  |  Data: {data_date}")
    print(DIVIDER)
    print()

    while True:
        # Step 1: search prompt
        term = input("Enter charity name to search (or Q to quit):\n> ").strip()
        if not term:
            continue
        if term.lower() in ("q", "quit"):
            break

        # Step 2: search
        results = df[df["NAME"].str.contains(term, case=False, na=False, regex=False)]
        if results.empty:
            print(f'\nNo matches found for "{term}". Try a shorter search term.\n')
            continue

        show_results(results, term)

        # Step 3: selection loop
        display_count = min(MAX_RESULTS, len(results))
        while True:
            sel = input("Enter number to select, S to search again, or Q to quit:\n> ").strip()
            if sel.lower() in ("q", "quit"):
                sys.exit(0)
            if sel.lower() == "s":
                break
            if sel.isdigit():
                n = int(sel)
                if 1 <= n <= display_count:
                    selected_row = results.iloc[n - 1]

                    # Find next empty row
                    next_row = find_next_row(xlsx_path)
                    if next_row is None:
                        print()
                        print("WARNING: All data rows in the spreadsheet appear to be filled.")
                        print(f"The spreadsheet supports {LAST_DATA_ROW - FIRST_DATA_ROW + 1} rows. You may need to add more rows manually,")
                        print("or start a new spreadsheet for additional entries.")
                        break

                    name, ein_fmt, address = show_confirmation(selected_row, next_row, xlsx_path.name)

                    # Step 4: confirm
                    while True:
                        confirm = input("Confirm? (Y/N):\n> ").strip().lower()
                        if confirm == "y":
                            ok = write_to_spreadsheet(xlsx_path, next_row, name, ein_fmt, address)
                            if ok:
                                print(f"\nDone. Written to row {next_row}.\n")
                            break
                        elif confirm == "n":
                            print("\nCancelled. Nothing was written.\n")
                            break
                        # anything else: re-display the prompt (loop)
                    break
                else:
                    print(f"Please enter a number between 1 and {display_count}.")
            else:
                print(f"Please enter a number between 1 and {display_count}, S to search again, or Q to quit.")


if __name__ == "__main__":
    main()
