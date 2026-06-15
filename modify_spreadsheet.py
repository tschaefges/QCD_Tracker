"""
One-time script: inserts the Charity Address column into QCD_Tracker_2026.xlsx.
Run once after setup, then this script can be kept for reference or discarded.

What it does:
  - Inserts a new column F (Charity Address) between EIN (E) and Amount (F→G)
  - Applies matching header and body formatting
  - Updates the SUM and REMAINING formulas that shifted from F to G
"""
from pathlib import Path
import glob as _glob

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

SCRIPT_DIR = Path(__file__).parent
SHEET_NAME = "QCD Transaction Log"

# Matched from existing spreadsheet (inspected with openpyxl)
HEADER_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL  = PatternFill(fill_type="solid", fgColor="FF2E75B6")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT       = Font(name="Arial", size=10)
BODY_FILL_GRAY  = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
BODY_FILL_WHITE = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
BODY_ALIGN      = Alignment(vertical="top", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="FFBFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def main():
    matches = list(SCRIPT_DIR.glob("QCD_Tracker*.xlsx"))
    if not matches:
        print("ERROR: No QCD_Tracker*.xlsx found in this folder.")
        print(f"  Folder checked: {SCRIPT_DIR}")
        return

    xlsx_path = matches[0]

    print(f"Opening {xlsx_path.name}...")
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except PermissionError:
        print(f"ERROR: {xlsx_path.name} is open in another program.")
        print("  Please close it and run this script again.")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found in {xlsx_path.name}.")
        return

    ws = wb[SHEET_NAME]

    # Safety check: if col F header already says "Charity Address", don't insert again
    current_f3 = ws.cell(row=3, column=6).value
    if current_f3 and "address" in str(current_f3).lower():
        print("Column F already appears to be Charity Address. Nothing changed.")
        return

    # Insert blank column at position 6 (F), shifting Amount and all columns right
    ws.insert_cols(6)

    # Header cell (row 3, col F)
    h = ws.cell(row=3, column=6)
    h.value     = "Charity Address"
    h.font      = HEADER_FONT
    h.fill      = HEADER_FILL
    h.alignment = HEADER_ALIGN

    # Body cells (rows 4-23) — even rows gray, odd rows white
    border = _thin_border()
    for row in range(4, 24):
        c = ws.cell(row=row, column=6)
        c.font      = BODY_FONT
        c.fill      = BODY_FILL_GRAY if row % 2 == 0 else BODY_FILL_WHITE
        c.border    = border
        c.alignment = BODY_ALIGN

    ws.column_dimensions["F"].width = 28

    # The Amount column shifted from F to G; explicitly correct the formulas
    # (openpyxl's insert_cols may or may not update formula references)
    ws["G24"].value = "=SUM(G4:G23)"
    ws["G25"].value = "=111000-G24"

    wb.save(xlsx_path)
    print(f"Done. {xlsx_path.name} updated.")
    print("  New column F: Charity Address")
    print("  Amount moved to column G; formulas in G24 and G25 updated.")
    print("  Open the spreadsheet in Excel to verify before using lookup_charity.exe.")


if __name__ == "__main__":
    main()
